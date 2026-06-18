#!/usr/bin/env python3
"""Build scene tags from scene-rule mappings and master-table rule tags."""

from __future__ import annotations

import argparse
import csv
from collections import Counter, defaultdict
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


DEFAULT_DIMENSIONS = [
    "attack_name",
    "attack_type",
    "malware",
    "threat_group",
    "vendor",
    "control",
    "mitre_tactics",
    "industries",
]


def clean(value: object) -> str:
    return str(value or "").strip()


def split_csv(value: str) -> list[str]:
    return [item.strip() for item in value.split(",") if item.strip()]


def style_ws(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for col in ws.columns:
        width = 10
        letter = col[0].column_letter
        for cell in col[:2000]:
            width = max(width, len(str(cell.value or "")) + 2)
        ws.column_dimensions[letter].width = min(width, 60)


def unique_preserving_order(values: Iterable[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        if value and value not in seen:
            seen.add(value)
            result.append(value)
    return result


def read_scene_mapping(path: Path, scene_uuid_col: str, scene_name_col: str, rule_uuid_col: str):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [clean(cell) for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    index = {header: i for i, header in enumerate(headers)}
    missing = [col for col in [scene_uuid_col, scene_name_col, rule_uuid_col] if col not in index]
    if missing:
        raise ValueError(f"Scene mapping missing columns: {missing}; found={headers}")

    scene_name: dict[str, str] = {}
    scene_rules: dict[str, list[str]] = defaultdict(list)
    for row in ws.iter_rows(min_row=2, values_only=True):
        scene_uuid = clean(row[index[scene_uuid_col]])
        name = clean(row[index[scene_name_col]])
        rule_uuid = clean(row[index[rule_uuid_col]])
        if not scene_uuid or not rule_uuid:
            continue
        scene_name[scene_uuid] = name
        scene_rules[scene_uuid].append(rule_uuid)

    for scene_uuid, rules in list(scene_rules.items()):
        scene_rules[scene_uuid] = unique_preserving_order(rules)
    return scene_name, scene_rules


def read_rule_tags(path: Path, needed_rules: set[str], dimensions: set[str]):
    rule_tags: dict[str, dict[str, set[tuple[str, str]]]] = defaultdict(lambda: defaultdict(set))
    rule_names: dict[str, str] = {}
    with path.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        required = {"rule_uuid", "rule_name", "tag_type", "tag_cn", "tag_en"}
        missing = required - set(reader.fieldnames or [])
        if missing:
            raise ValueError(f"Rule tag CSV missing columns: {sorted(missing)}")
        for record in reader:
            rule_uuid = clean(record.get("rule_uuid"))
            if rule_uuid not in needed_rules:
                continue
            dimension = clean(record.get("tag_type"))
            if dimension not in dimensions:
                continue
            tag_cn = clean(record.get("tag_cn")) or clean(record.get("raw_value"))
            tag_en = clean(record.get("tag_en"))
            if not tag_cn and not tag_en:
                continue
            rule_names[rule_uuid] = clean(record.get("rule_name"))
            rule_tags[rule_uuid][dimension].add((tag_cn, tag_en))
    return rule_tags, rule_names


def write_outputs(
    *,
    scene_name: dict[str, str],
    scene_rules: dict[str, list[str]],
    rule_tags: dict[str, dict[str, set[tuple[str, str]]]],
    rule_names: dict[str, str],
    dimensions: list[str],
    threshold: float,
    out_dir: Path,
    prefix: str,
) -> tuple[Path, Path, Counter]:
    out_dir.mkdir(parents=True, exist_ok=True)
    ratio_path = out_dir / f"{prefix}_ratio.xlsx"
    final_path = out_dir / f"{prefix}_ge80.xlsx"

    summary_rows: list[list[object]] = []
    ratio_rows: list[list[object]] = []
    final_by_dim: dict[str, list[list[object]]] = defaultdict(list)
    scene_rule_map: list[list[object]] = []
    missing_rows: list[list[object]] = []
    stats = Counter()

    for scene_uuid in sorted(scene_rules, key=lambda item: (scene_name.get(item, ""), item)):
        rules = scene_rules[scene_uuid]
        total = len(rules)
        matched = sum(1 for rule_uuid in rules if rule_uuid in rule_tags)
        missing = [rule_uuid for rule_uuid in rules if rule_uuid not in rule_tags]
        summary_rows.append([scene_uuid, scene_name.get(scene_uuid, ""), total, matched, len(missing), matched / total if total else 0])
        for rule_uuid in missing:
            missing_rows.append([scene_uuid, scene_name.get(scene_uuid, ""), rule_uuid])
        for rule_uuid in rules:
            scene_rule_map.append([scene_uuid, scene_name.get(scene_uuid, ""), rule_uuid, rule_names.get(rule_uuid, ""), "Y" if rule_uuid in rule_tags else "N"])

        for dimension in dimensions:
            tag_to_rules: dict[tuple[str, str], set[str]] = defaultdict(set)
            for rule_uuid in rules:
                for tag in rule_tags.get(rule_uuid, {}).get(dimension, set()):
                    tag_to_rules[tag].add(rule_uuid)
            for (tag_cn, tag_en), support_rules in sorted(tag_to_rules.items(), key=lambda item: (-len(item[1]), item[0][0], item[0][1])):
                support_count = len(support_rules)
                ratio = support_count / total if total else 0
                ratio_rows.append([
                    scene_uuid,
                    scene_name.get(scene_uuid, ""),
                    dimension,
                    total,
                    matched,
                    tag_cn,
                    tag_en,
                    support_count,
                    ratio,
                    " | ".join(sorted(support_rules)),
                ])
                if ratio >= threshold:
                    final_by_dim[dimension].append([scene_uuid, scene_name.get(scene_uuid, ""), tag_cn, tag_en, support_count, ratio])
                    stats[f"final_{dimension}"] += 1

    stat_pairs: list[tuple[str, object]] = [
        ("scene_count", len(scene_rules)),
        ("scene_rule_edges", sum(len(rules) for rules in scene_rules.values())),
        ("unique_rule_count", len({rule_uuid for rules in scene_rules.values() for rule_uuid in rules})),
        ("matched_unique_rule_count", len(rule_tags)),
        ("missing_unique_rule_count", len({rule_uuid for rules in scene_rules.values() for rule_uuid in rules} - set(rule_tags))),
        ("ratio_threshold", threshold),
    ]
    for dimension in dimensions:
        stat_pairs.append((f"final_tag_rows_{dimension}", len(final_by_dim.get(dimension, []))))

    wb_ratio = Workbook()
    ws = wb_ratio.active
    ws.title = "scene_summary"
    ws.append(["scene_uuid", "scene_name", "scene_rule_count", "matched_rule_count", "missing_rule_count", "matched_rule_ratio"])
    for row in summary_rows:
        ws.append(row)
    style_ws(ws)
    for cell in ws.iter_rows(min_row=2, min_col=6, max_col=6):
        cell[0].number_format = "0.00%"

    ws = wb_ratio.create_sheet("scene_tag_ratio")
    ws.append(["scene_uuid", "scene_name", "dimension", "scene_rule_count", "matched_rule_count", "tag_cn", "tag_en", "support_rule_count", "ratio_in_scene_rules", "support_rule_uuids"])
    for row in ratio_rows:
        ws.append(row)
    style_ws(ws)
    for cell in ws.iter_rows(min_row=2, min_col=9, max_col=9):
        cell[0].number_format = "0.00%"

    ws = wb_ratio.create_sheet("scene_rule_map")
    ws.append(["scene_uuid", "scene_name", "rule_uuid", "rule_name", "rule_found_in_master_tags"])
    for row in scene_rule_map:
        ws.append(row)
    style_ws(ws)

    ws = wb_ratio.create_sheet("missing_rules")
    ws.append(["scene_uuid", "scene_name", "missing_rule_uuid"])
    for row in missing_rows:
        ws.append(row)
    style_ws(ws)

    ws = wb_ratio.create_sheet("stats")
    ws.append(["metric", "value"])
    for key, value in stat_pairs:
        ws.append([key, value])
    style_ws(ws)
    wb_ratio.save(ratio_path)

    wb_final = Workbook()
    wb_final.remove(wb_final.active)
    ws = wb_final.create_sheet("overview")
    ws.append(["metric", "value"])
    for key, value in stat_pairs:
        ws.append([key, value])
    style_ws(ws)
    for dimension in dimensions:
        ws = wb_final.create_sheet(dimension)
        ws.append(["uuid", "场景名字", "cntag", "entag", "support_rule_count", "ratio"])
        for row in final_by_dim.get(dimension, []):
            ws.append(row)
        style_ws(ws)
        for cell in ws.iter_rows(min_row=2, min_col=6, max_col=6):
            cell[0].number_format = "0.00%"
    wb_final.save(final_path)
    return ratio_path, final_path, stats


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--scene-xlsx", required=True, type=Path)
    parser.add_argument("--rule-tags-csv", required=True, type=Path)
    parser.add_argument("--out-dir", required=True, type=Path)
    parser.add_argument("--prefix", default="scene_tags")
    parser.add_argument("--threshold", type=float, default=0.8)
    parser.add_argument("--dimensions", default=",".join(DEFAULT_DIMENSIONS), help="Comma-separated tag dimensions.")
    parser.add_argument("--scene-uuid-col", default="UUID")
    parser.add_argument("--scene-name-col", default="name")
    parser.add_argument("--rule-uuid-col", default="sim_actions")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    dimensions = split_csv(args.dimensions)
    scene_name, scene_rules = read_scene_mapping(args.scene_xlsx, args.scene_uuid_col, args.scene_name_col, args.rule_uuid_col)
    needed_rules = {rule_uuid for rules in scene_rules.values() for rule_uuid in rules}
    rule_tags, rule_names = read_rule_tags(args.rule_tags_csv, needed_rules, set(dimensions))
    ratio_path, final_path, stats = write_outputs(
        scene_name=scene_name,
        scene_rules=scene_rules,
        rule_tags=rule_tags,
        rule_names=rule_names,
        dimensions=dimensions,
        threshold=args.threshold,
        out_dir=args.out_dir,
        prefix=args.prefix,
    )
    print(f"scene_count={len(scene_rules)}")
    print(f"scene_rule_edges={sum(len(rules) for rules in scene_rules.values())}")
    print(f"unique_rule_count={len(needed_rules)}")
    print(f"matched_unique_rule_count={len(rule_tags)}")
    print(f"missing_unique_rule_count={len(needed_rules - set(rule_tags))}")
    for dimension in dimensions:
        print(f"final_tag_rows_{dimension}={stats.get(f'final_{dimension}', 0)}")
    print(f"ratio_workbook={ratio_path}")
    print(f"final_workbook={final_path}")


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import uuid
from pathlib import Path
from typing import Dict, Iterable, List, Optional

CONTROL_ALIASES = {
    "control": "control",
    "安全控制设备": "control",
    "安全控制手背": "control",
    "安全控制设备(control)": "control",
    "安全控制设备（control）": "control",
}


def clean(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text.lower() == "nan" else text


def normalize_header(value) -> str:
    return re.sub(r"[\s_\-/:（）()【】\[\]．。]+", "", str(value or "").strip()).lower()


def find_col(columns: Iterable[str], aliases: Iterable[str]) -> Optional[str]:
    wanted = {normalize_header(alias) for alias in aliases}
    for column in columns:
        if normalize_header(column) in wanted:
            return column
    return None


def normalize_tag_type(value: str | None) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    compact = text.lower().replace("（", "(").replace("）", ")")
    compact = re.sub(r"\s+", "", compact)
    return CONTROL_ALIASES.get(compact, text.lower())


def read_language_map(path: Optional[Path], allowed_uuids: Optional[set[str]] = None) -> Dict[str, Dict[str, str]]:
    if not path:
        return {}
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    language: Dict[str, Dict[str, str]] = {}

    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        try:
            headers = [clean(x) for x in next(rows)]
        except StopIteration:
            continue

        idx = {header: i for i, header in enumerate(headers)}
        colmap = {
            "uuid": find_col(headers, ["uuid", "name", "规则uuid", "id"]),
            "rule_name": find_col(headers, ["cn_name", "rule_name", "name.1", "规则名称", "namecn"]),
            "rule_name_en": find_col(headers, ["en_name", "rule_name_en", "name_en", "英文名称"]),
            "description_cn": find_col(headers, ["cn_desc", "description_cn", "desc", "描述", "规则描述"]),
            "description_en": find_col(headers, ["en_desc", "description_en", "desc_en", "英文描述"]),
            "note_cn": find_col(headers, ["cn_notes", "note_cn", "notes", "备注", "说明"]),
            "note_en": find_col(headers, ["en_notes", "note_en", "notes_en", "cn_notes_en", "英文备注"]),
        }
        if not colmap["uuid"]:
            continue

        def get(row, key: str) -> str:
            column = colmap[key]
            if not column or column not in idx or idx[column] >= len(row):
                return ""
            return clean(row[idx[column]])

        for row in rows:
            rule_uuid = get(row, "uuid").lower()
            if not rule_uuid:
                continue
            if allowed_uuids and rule_uuid not in allowed_uuids:
                continue

            previous = language.get(rule_uuid, {})
            merged = dict(previous)
            # Non-empty override only. This protects Actions metadata from being blanked by Email sheets.
            for key in ["rule_name", "rule_name_en", "description_cn", "description_en", "note_cn", "note_en"]:
                value = get(row, key)
                if value:
                    merged[key] = value
                else:
                    merged.setdefault(key, "")
            language[rule_uuid] = merged

    return language


def read_classified(path: Path, language: Dict[str, Dict[str, str]]) -> tuple[list[list[str]], list[list[str]], list[str]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    modified: list[list[str]] = []
    modified_seen: set[str] = set()
    tags: list[list[str]] = []
    tag_types: set[str] = set()

    def add_modified(rule_uuid: str) -> None:
        if not rule_uuid or rule_uuid in modified_seen:
            return
        modified_seen.add(rule_uuid)
        meta = language.get(rule_uuid.lower(), {})
        modified.append([
            rule_uuid,
            meta.get("rule_name", ""),
            meta.get("rule_name_en", ""),
            meta.get("description_cn", ""),
            meta.get("description_en", ""),
            meta.get("note_cn", ""),
            meta.get("note_en", ""),
        ])

    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        try:
            headers = [clean(x) for x in next(rows)]
        except StopIteration:
            continue
        if "uuid" not in headers:
            continue
        idx = {header: i for i, header in enumerate(headers)}

        if ws.title == "modified_rules":
            for row in rows:
                rule_uuid = clean(row[idx["uuid"]] if idx["uuid"] < len(row) else "")
                add_modified(rule_uuid)
            continue

        tag_type_col = "tag_type" if "tag_type" in idx else None
        tag_cn_col = "tag_cn" if "tag_cn" in idx else (headers[2] if len(headers) >= 3 else None)
        tag_en_col = "tag_en" if "tag_en" in idx else (headers[3] if len(headers) >= 4 else None)

        for row in rows:
            rule_uuid = clean(row[idx["uuid"]] if idx["uuid"] < len(row) else "")
            if not rule_uuid:
                continue
            raw_type = clean(row[idx[tag_type_col]]) if tag_type_col and idx[tag_type_col] < len(row) else ws.title
            tag_type = normalize_tag_type(raw_type or ws.title)
            tag_cn = clean(row[idx[tag_cn_col]]) if tag_cn_col and idx[tag_cn_col] < len(row) else ""
            tag_en = clean(row[idx[tag_en_col]]) if tag_en_col and idx[tag_en_col] < len(row) else ""
            if not tag_type or (not tag_cn and not tag_en):
                continue
            tags.append([rule_uuid, tag_type, tag_cn, tag_en])
            tag_types.add(tag_type)
            add_modified(rule_uuid)

    return modified, tags, sorted(tag_types)


def collect_modified_uuids(path: Path) -> set[str]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    found: set[str] = set()
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        try:
            headers = [clean(x) for x in next(rows)]
        except StopIteration:
            continue
        if ws.title != "modified_rules" or "uuid" not in headers:
            continue
        idx = {header: i for i, header in enumerate(headers)}
        for row in rows:
            rule_uuid = clean(row[idx["uuid"]] if idx["uuid"] < len(row) else "")
            if rule_uuid:
                found.add(rule_uuid.lower())
    return found


def write_csvs(workdir: Path, modified: list[list[str]], tags: list[list[str]], tag_types: list[str]) -> None:
    with (workdir / "modified.csv").open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["rule_uuid", "rule_name", "rule_name_en", "description_cn", "description_en", "note_cn", "note_en"])
        writer.writerows(modified)
    with (workdir / "tags.csv").open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["rule_uuid", "tag_type", "tag_cn", "tag_en"])
        writer.writerows(tags)
    with (workdir / "tag_types.csv").open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["tag_type"])
        for tag_type in tag_types:
            writer.writerow([tag_type])


def sql_literal(value: str) -> str:
    return value.replace("'", "''")


def build_sql(args, new_id: str) -> str:
    version_name = sql_literal(args.version_name)
    description = sql_literal(f"Imported classified tags from {Path(args.classified).name}" + (f" with language file {Path(args.language).name}" if args.language else ""))
    base_id = args.base_version_id
    rule_set = args.rule_set.upper()
    return f"""
\\set ON_ERROR_STOP on
BEGIN;
CREATE TEMP TABLE tmp_modified(rule_uuid text, rule_name text, rule_name_en text, description_cn text, description_en text, note_cn text, note_en text);
CREATE TEMP TABLE tmp_tags(rule_uuid text, tag_type text, tag_cn text, tag_en text);
CREATE TEMP TABLE tmp_tag_types(tag_type text);
\\copy tmp_modified FROM '{args.remote_dir}/modified.csv' WITH (FORMAT csv, HEADER true)
\\copy tmp_tags FROM '{args.remote_dir}/tags.csv' WITH (FORMAT csv, HEADER true)
\\copy tmp_tag_types FROM '{args.remote_dir}/tag_types.csv' WITH (FORMAT csv, HEADER true)

DO $$
BEGIN
    IF EXISTS (SELECT 1 FROM master_table_version WHERE version_name = '{version_name}') THEN
        RAISE EXCEPTION 'target version already exists: {version_name}';
    END IF;
END $$;

INSERT INTO master_table_version(id, version_name, description, is_base, base_version_id, created_at)
VALUES ('{new_id}'::uuid, '{version_name}', '{description}', false, '{base_id}'::uuid, now());

INSERT INTO master_table_entry(id, version_id, rule_uuid, rule_name, rule_set, created_at, rule_name_en, description_cn, description_en, note_cn, note_en, merged_at)
SELECT regexp_replace(md5('{new_id}' || ':' || e.id::text), '^(.{{8}})(.{{4}})(.{{4}})(.{{4}})(.{{12}})$', '\\1-\\2-\\3-\\4-\\5')::uuid,
       '{new_id}'::uuid, e.rule_uuid, e.rule_name, e.rule_set, e.created_at, e.rule_name_en, e.description_cn, e.description_en, e.note_cn, e.note_en, e.merged_at
FROM master_table_entry e
WHERE e.version_id = '{base_id}'::uuid;

INSERT INTO master_table_entry(id, version_id, rule_uuid, rule_name, rule_set, created_at, rule_name_en, description_cn, description_en, note_cn, note_en, merged_at)
SELECT regexp_replace(md5('{new_id}' || ':new:' || m.rule_uuid), '^(.{{8}})(.{{4}})(.{{4}})(.{{4}})(.{{12}})$', '\\1-\\2-\\3-\\4-\\5')::uuid,
       '{new_id}'::uuid, m.rule_uuid, COALESCE(NULLIF(m.rule_name,''), m.rule_uuid),
       '{rule_set}', now(), NULLIF(m.rule_name_en,''), NULLIF(m.description_cn,''), NULLIF(m.description_en,''), NULLIF(m.note_cn,''), NULLIF(m.note_en,''), now()
FROM tmp_modified m
WHERE NOT EXISTS (
    SELECT 1 FROM master_table_entry e WHERE e.version_id = '{new_id}'::uuid AND e.rule_uuid = m.rule_uuid
);

UPDATE master_table_entry e
SET rule_name = COALESCE(NULLIF(m.rule_name,''), e.rule_name),
    rule_name_en = COALESCE(NULLIF(m.rule_name_en,''), e.rule_name_en),
    description_cn = COALESCE(NULLIF(m.description_cn,''), e.description_cn),
    description_en = COALESCE(NULLIF(m.description_en,''), e.description_en),
    note_cn = COALESCE(NULLIF(m.note_cn,''), e.note_cn),
    note_en = COALESCE(NULLIF(m.note_en,''), e.note_en),
    rule_set = '{rule_set}',
    merged_at = now()
FROM tmp_modified m
WHERE e.version_id = '{new_id}'::uuid AND e.rule_uuid = m.rule_uuid;

WITH entry_map AS (
    SELECT oe.id AS old_entry_id, ne.id AS new_entry_id, oe.rule_uuid
    FROM master_table_entry oe
    JOIN master_table_entry ne
      ON ne.version_id = '{new_id}'::uuid
     AND ne.rule_uuid = oe.rule_uuid
    WHERE oe.version_id = '{base_id}'::uuid
)
INSERT INTO master_table_tag(id, entry_id, tag_dict_id, raw_value, tag_type, mapping_status)
SELECT regexp_replace(md5('{new_id}' || ':' || t.id::text), '^(.{{8}})(.{{4}})(.{{4}})(.{{4}})(.{{12}})$', '\\1-\\2-\\3-\\4-\\5')::uuid,
       em.new_entry_id, t.tag_dict_id, t.raw_value, t.tag_type, t.mapping_status
FROM entry_map em
JOIN master_table_tag t ON t.entry_id = em.old_entry_id
WHERE NOT (
    EXISTS (SELECT 1 FROM tmp_modified m WHERE m.rule_uuid = em.rule_uuid)
    AND EXISTS (SELECT 1 FROM tmp_tag_types tt WHERE tt.tag_type = t.tag_type)
);

WITH active_versions AS (
    SELECT id FROM dictionary_version WHERE is_mapping_active = true
    UNION ALL
    SELECT id FROM dictionary_version
    WHERE NOT EXISTS (SELECT 1 FROM dictionary_version WHERE is_mapping_active = true)
      AND is_active = true
), dict_candidates AS (
    SELECT td.id, td.tag_type, td.tag_cn AS val, td.created_at
    FROM tag_dictionary td JOIN active_versions av ON av.id = td.version_id
    WHERE td.is_active = true AND td.review_status = 'APPROVED' AND td.tag_cn IS NOT NULL AND td.tag_cn <> ''
    UNION ALL
    SELECT td.id, td.tag_type, td.tag_en AS val, td.created_at
    FROM tag_dictionary td JOIN active_versions av ON av.id = td.version_id
    WHERE td.is_active = true AND td.review_status = 'APPROVED' AND td.tag_en IS NOT NULL AND td.tag_en <> ''
), dict_lookup AS (
    SELECT DISTINCT ON (tag_type, val) tag_type, val, id
    FROM dict_candidates
    ORDER BY tag_type, val, created_at NULLS LAST, id
), prepared AS (
    SELECT e.id AS entry_id,
           COALESCE(d1.id, d2.id) AS tag_dict_id,
           COALESCE(NULLIF(t.tag_cn,''), NULLIF(t.tag_en,'')) AS raw_value,
           t.tag_type,
           CASE WHEN COALESCE(d1.id, d2.id) IS NULL THEN 'UNMAPPED' ELSE 'MANUAL' END AS mapping_status
    FROM tmp_tags t
    JOIN master_table_entry e ON e.version_id = '{new_id}'::uuid AND e.rule_uuid = t.rule_uuid
    LEFT JOIN dict_lookup d1 ON d1.tag_type = t.tag_type AND d1.val = NULLIF(t.tag_cn,'')
    LEFT JOIN dict_lookup d2 ON d2.tag_type = t.tag_type AND d2.val = NULLIF(t.tag_en,'')
    WHERE COALESCE(NULLIF(t.tag_cn,''), NULLIF(t.tag_en,'')) IS NOT NULL
), deduped AS (
    SELECT DISTINCT ON (entry_id, COALESCE(tag_dict_id::text,''), tag_type, raw_value)
           entry_id, tag_dict_id, raw_value, tag_type, mapping_status
    FROM prepared
    ORDER BY entry_id, COALESCE(tag_dict_id::text,''), tag_type, raw_value
)
INSERT INTO master_table_tag(id, entry_id, tag_dict_id, raw_value, tag_type, mapping_status)
SELECT regexp_replace(md5('{new_id}' || ':tag:' || entry_id::text || ':' || COALESCE(tag_dict_id::text,'') || ':' || tag_type || ':' || raw_value), '^(.{{8}})(.{{4}})(.{{4}})(.{{4}})(.{{12}})$', '\\1-\\2-\\3-\\4-\\5')::uuid,
       entry_id, tag_dict_id, raw_value, tag_type, mapping_status
FROM deduped;

COMMIT;

SELECT 'new_version_id', '{new_id}';
SELECT 'modified_uuids', count(*) FROM tmp_modified;
SELECT 'input_tag_rows', count(*) FROM tmp_tags;
SELECT 'version_entries', count(*) FROM master_table_entry WHERE version_id = '{new_id}'::uuid;
SELECT 'version_distinct_entries', count(distinct rule_uuid) FROM master_table_entry WHERE version_id = '{new_id}'::uuid;
SELECT 'version_tags', count(*) FROM master_table_tag t JOIN master_table_entry e ON e.id=t.entry_id WHERE e.version_id = '{new_id}'::uuid;
SELECT 'still_uuid_name', count(*) FROM master_table_entry e JOIN tmp_modified m ON m.rule_uuid=e.rule_uuid WHERE e.version_id = '{new_id}'::uuid AND (e.rule_name IS NULL OR e.rule_name='' OR e.rule_name=e.rule_uuid);
"""


def run(cmd: List[str], env: Optional[dict] = None) -> None:
    print("+", " ".join(cmd), file=sys.stderr)
    subprocess.run(cmd, check=True, env=env)


def upload_and_execute(args, workdir: Path) -> None:
    sshpass = shutil.which("sshpass")
    if not sshpass:
        raise SystemExit("sshpass is required for password SSH. Install it or upload/execute manually.")
    if not args.ssh_password:
        raise SystemExit("Missing --ssh-password or TAGSYS_SSH_PASSWORD.")
    if not args.db_password:
        raise SystemExit("Missing --db-password or TAGSYS_DB_PASSWORD.")

    ssh_base = [
        sshpass, "-p", args.ssh_password, "ssh", "-T",
        "-o", "StrictHostKeyChecking=no",
        "-o", "PreferredAuthentications=password",
        "-o", "PubkeyAuthentication=no",
        f"{args.ssh_user}@{args.host}",
    ]
    scp_base = [
        sshpass, "-p", args.ssh_password, "scp",
        "-o", "StrictHostKeyChecking=no",
        "-o", "PreferredAuthentications=password",
        "-o", "PubkeyAuthentication=no",
    ]

    run(ssh_base + [f"rm -rf {args.remote_dir} && mkdir -p {args.remote_dir}"])
    upload_files = [str(workdir / name) for name in ["modified.csv", "tags.csv", "tag_types.csv", "merge.sql"]]
    run(scp_base + upload_files + [f"{args.ssh_user}@{args.host}:{args.remote_dir}/"])

    remote_cmd = (
        f"export PGPASSWORD={args.db_password}; "
        f"psql -h 127.0.0.1 -U {args.db_user} -d {args.db_name} -f {args.remote_dir}/merge.sql"
    )
    run(ssh_base + [remote_cmd])


def main() -> None:
    parser = argparse.ArgumentParser(description="Fast merge classified tag workbook into online master table.")
    parser.add_argument("--classified", required=True)
    parser.add_argument("--language")
    parser.add_argument("--rule-set", required=True, choices=["VALIDATION", "DETECTION"])
    parser.add_argument("--base-version-id", required=True)
    parser.add_argument("--version-name", required=True)
    parser.add_argument("--host", default="192.168.10.89")
    parser.add_argument("--ssh-user", default="dx")
    parser.add_argument("--ssh-password", default=os.environ.get("TAGSYS_SSH_PASSWORD"))
    parser.add_argument("--db-name", default="tag_system")
    parser.add_argument("--db-user", default="tagapp")
    parser.add_argument("--db-password", default=os.environ.get("TAGSYS_DB_PASSWORD"))
    parser.add_argument("--remote-dir", default="/tmp/master_classified_merge")
    parser.add_argument("--workdir")
    parser.add_argument("--execute", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    classified = Path(args.classified).expanduser().resolve()
    language = Path(args.language).expanduser().resolve() if args.language else None
    if not classified.exists():
        raise SystemExit(f"classified workbook not found: {classified}")
    if language and not language.exists():
        raise SystemExit(f"language workbook not found: {language}")

    modified_uuids = collect_modified_uuids(classified)
    language_map = read_language_map(language, modified_uuids)
    modified, tags, tag_types = read_classified(classified, language_map)
    missing_names = [row[0] for row in modified if not row[1]]

    new_id = str(uuid.uuid4())
    workdir = Path(args.workdir).expanduser().resolve() if args.workdir else Path(tempfile.mkdtemp(prefix="master_classified_merge_"))
    workdir.mkdir(parents=True, exist_ok=True)
    write_csvs(workdir, modified, tags, tag_types)
    (workdir / "merge.sql").write_text(build_sql(args, new_id), encoding="utf-8")

    summary = {
        "workdir": str(workdir),
        "new_version_id": new_id,
        "classified": str(classified),
        "language": str(language) if language else None,
        "rule_set": args.rule_set,
        "base_version_id": args.base_version_id,
        "version_name": args.version_name,
        "modified_uuids": len(modified),
        "input_tag_rows": len(tags),
        "tag_types": tag_types,
        "language_matched_uuids": len(language_map),
        "missing_rule_name_count_before_merge": len(missing_names),
        "missing_rule_name_samples": missing_names[:20],
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))

    if args.dry_run and not args.execute:
        return
    if args.execute:
        upload_and_execute(args, workdir)


if __name__ == "__main__":
    main()
